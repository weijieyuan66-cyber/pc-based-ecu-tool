"""
validation/matrix_loader.py
----------------------------
ExpectationMatrixLoader — loads an ExpectationSpec from an Excel workbook
at runtime, without requiring any code changes.

Workbook structure
------------------
  Sheet "Messages"  (required)
    Columns (order-independent, matched by header name):
      ecu_name                      str    — informational only
      direction                     str    — rx_to_ecu / tx_from_ecu (informational)
      message_name                  str    — becomes ExpectedMessageSpec.label
      frame_id                      str    — hex (0x100) or decimal int
      pgn                           str    — optional hex / decimal (informational)
      required                      str    — yes/true/1 → True, else False
      expected_source_address       str    — hex or decimal; blank → None (skip check)
      expected_destination_address  str    — reserved / informational only
      expected_cycle_time_ms        float  — blank → None (skip check)
      cycle_tolerance_ms            float  — absolute ms; converted to % for validator
      timeout_ms                    float  — reserved / informational only
      expected_dlc                  int    — informational only (not validated in MVP)
      description                   str    — informational only

  Sheet "FieldConstraints"  (required)
    Columns (order-independent):
      message_name     str   — must match a message_name in Messages sheet
      signal_name      str   — becomes ExpectedFieldConstraint.field_name
      constraint_type  str   — "fixed_value" (only type supported in MVP)
      expected_value   str   — hex (0x3E) or decimal int
      byte_index       int   — zero-based byte position in the CAN frame payload
      tolerance        float — reserved / not used by fixed_value in MVP
      description      str   — informational only

  Sheet "Metadata"  (optional)
    Columns: key, value
      Recognised keys:
        scenario_name  — sets ExpectationSpec.scenario_name

Usage
-----
  from validation.matrix_loader import ExpectationMatrixLoader

  spec = ExpectationMatrixLoader.load("path/to/matrix.xlsx")
  # → ExpectationSpec ready to pass to ExpectationValidator

  spec = ExpectationMatrixLoader.load_from_config(config_dict)
  # → ExpectationSpec if config["expectation_matrix_path"] is set, else None
"""

from __future__ import annotations

import logging
from typing import Dict, List, Optional, Tuple

from validation.results import DeviationType  # noqa: F401 (for type safety)
from validation.specs import (
    ExpectedFieldConstraint,
    ExpectedMessageSpec,
    ExpectationSpec,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Internal parsing helpers
# ---------------------------------------------------------------------------

def _parse_numeric_to_int(value) -> Optional[int]:
    """Return an int from a cell value that may be a hex string, int, or float."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(s, 0)  # handles "0x100", "256", "100" etc.
    except ValueError:
        logger.warning("Could not parse integer/hex value %r — skipping.", s)
        return None


def _parse_bool(value) -> bool:
    """Parse yes/true/1 → True, everything else → False."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ("yes", "true", "1")


def _parse_float(value) -> Optional[float]:
    """Return a float or None for blank/None cells."""
    if value is None or (isinstance(value, str) and str(value).strip() == ""):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        logger.warning("Could not parse float value %r — skipping.", value)
        return None


def _row_to_dict(headers: List[str], row) -> Dict[str, object]:
    """Zip header names with cell values, normalising header names to lower-case."""
    return {
        headers[i].strip().lower(): (cell.value if hasattr(cell, "value") else cell)
        for i, cell in enumerate(row)
        if i < len(headers)
    }


# ---------------------------------------------------------------------------
# Public loader
# ---------------------------------------------------------------------------

class ExpectationMatrixLoader:
    """
    Loads an ExpectationSpec from an Excel workbook (.xlsx / .xls).

    All parsing is done in pure Python via openpyxl.  No UI or
    transport-layer imports.
    """

    # ------------------------------------------------------------------ #
    # Public API                                                           #
    # ------------------------------------------------------------------ #

    @staticmethod
    def load(path: str) -> ExpectationSpec:
        """
        Load an ExpectationSpec from the given Excel workbook path.

        Parameters
        ----------
        path : str
            File-system path to an .xlsx workbook.

        Returns
        -------
        ExpectationSpec

        Raises
        ------
        ImportError
            If openpyxl is not installed.
        FileNotFoundError / ValueError
            If the file cannot be opened or required sheets are missing.
        """
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required to load Expectation Matrix files.  "
                "Install it with:  pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        logger.info("Loading Expectation Matrix from: %s", path)

        scenario_name = ExpectationMatrixLoader._read_metadata(wb)
        constraints_map = ExpectationMatrixLoader._read_field_constraints(wb)
        messages = ExpectationMatrixLoader._read_messages(wb, constraints_map)

        spec = ExpectationSpec(
            scenario_name=scenario_name,
            messages=messages,
        )
        logger.info(
            "Expectation Matrix loaded: scenario='%s'  messages=%d",
            spec.scenario_name, len(spec.messages),
        )
        return spec

    @staticmethod
    def load_from_config(config: dict) -> Optional[ExpectationSpec]:
        """
        Load an ExpectationSpec from the path stored in config, if any.

        Parameters
        ----------
        config : dict
            Application config dict.  The key ``expectation_matrix_path``
            is checked; if absent or null/empty, returns None.

        Returns
        -------
        ExpectationSpec or None
        """
        path = config.get("expectation_matrix_path") or ""
        if not path:
            return None
        return ExpectationMatrixLoader.load(str(path))

    # ------------------------------------------------------------------ #
    # Sheet readers                                                        #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _read_metadata(wb) -> str:
        """Return scenario_name from the Metadata sheet, or 'loaded_matrix'."""
        if "Metadata" not in wb.sheetnames:
            return "loaded_matrix"

        ws = wb["Metadata"]
        rows = list(ws.iter_rows(min_row=1))
        if not rows:
            return "loaded_matrix"

        # Find header row (should be row 1: key, value)
        headers = [
            (cell.value or "").strip().lower()
            for cell in rows[0]
        ]
        if "key" not in headers or "value" not in headers:
            return "loaded_matrix"

        key_idx = headers.index("key")
        val_idx = headers.index("value")

        for row in rows[1:]:
            if len(row) <= max(key_idx, val_idx):
                continue
            k = (row[key_idx].value or "")
            v = (row[val_idx].value or "")
            if str(k).strip().lower() == "scenario_name" and v:
                return str(v).strip()

        return "loaded_matrix"

    @staticmethod
    def _read_field_constraints(
        wb,
    ) -> Dict[str, List[ExpectedFieldConstraint]]:
        """
        Read the FieldConstraints sheet and return a dict mapping
        message_name → list of ExpectedFieldConstraint.
        """
        result: Dict[str, List[ExpectedFieldConstraint]] = {}

        if "FieldConstraints" not in wb.sheetnames:
            logger.debug("No FieldConstraints sheet found — no field checks loaded.")
            return result

        ws = wb["FieldConstraints"]
        rows = list(ws.iter_rows(min_row=1))
        if len(rows) < 2:
            return result

        headers = [str(cell.value or "").strip().lower() for cell in rows[0]]

        for row in rows[1:]:
            d = _row_to_dict(headers, row)
            msg_name = str(d.get("message_name") or "").strip()
            if not msg_name:
                continue

            constraint_type = str(d.get("constraint_type") or "").strip().lower()
            if constraint_type != "fixed_value":
                logger.debug(
                    "Unsupported constraint_type '%s' for message '%s' — skipped.",
                    constraint_type, msg_name,
                )
                continue

            expected_val = _parse_numeric_to_int(d.get("expected_value"))
            if expected_val is None:
                logger.warning(
                    "FieldConstraints row for '%s': missing expected_value — skipped.",
                    msg_name,
                )
                continue

            byte_index_raw = d.get("byte_index")
            # byte_index defaults to 0 when the column is blank or absent.
            # This is the most common case (single-byte constraints on byte 0).
            # Provide a non-zero value in the sheet to target a different byte.
            byte_index = int(byte_index_raw) if byte_index_raw not in (None, "") else 0
            signal_name = str(d.get("signal_name") or "field").strip()

            fc = ExpectedFieldConstraint(
                field_index=byte_index,
                expected_value=expected_val,
                field_name=signal_name,
            )
            result.setdefault(msg_name, []).append(fc)

        return result

    @staticmethod
    def _read_messages(
        wb,
        constraints_map: Dict[str, List[ExpectedFieldConstraint]],
    ) -> List[ExpectedMessageSpec]:
        """
        Read the Messages sheet and return a list of ExpectedMessageSpec.
        Field constraints are attached from constraints_map.
        """
        if "Messages" not in wb.sheetnames:
            raise ValueError(
                "Excel workbook is missing the required 'Messages' sheet."
            )

        ws = wb["Messages"]
        rows = list(ws.iter_rows(min_row=1))
        if len(rows) < 2:
            return []

        headers = [str(cell.value or "").strip().lower() for cell in rows[0]]
        messages: List[ExpectedMessageSpec] = []

        for row_idx, row in enumerate(rows[1:], start=2):
            d = _row_to_dict(headers, row)

            msg_name = str(d.get("message_name") or "").strip()
            if not msg_name:
                continue  # skip blank rows

            frame_id = _parse_numeric_to_int(d.get("frame_id"))
            if frame_id is None:
                logger.warning(
                    "Messages row %d ('%s'): missing or invalid frame_id — skipped.",
                    row_idx, msg_name,
                )
                continue

            # A message is required only when the 'required' cell is explicitly
            # set to yes/true/1.  Blank or absent defaults to False (not required).
            required = _parse_bool(d.get("required"))
            sa_raw = d.get("expected_source_address")
            expected_sa = _parse_numeric_to_int(sa_raw)

            cycle_ms = _parse_float(d.get("expected_cycle_time_ms"))
            tolerance_ms = _parse_float(d.get("cycle_tolerance_ms"))

            # Convert absolute-ms tolerance to a percentage for the validator.
            cycle_tol_pct = 20.0  # default
            if cycle_ms and tolerance_ms and cycle_ms > 0:
                cycle_tol_pct = (tolerance_ms / cycle_ms) * 100.0

            fc_list: Tuple[ExpectedFieldConstraint, ...] = tuple(
                constraints_map.get(msg_name, [])
            )

            spec = ExpectedMessageSpec(
                message_id=frame_id,
                label=msg_name,
                required=required,
                expected_source_address=expected_sa,
                expected_cycle_time_ms=cycle_ms,
                cycle_time_tolerance_pct=cycle_tol_pct,
                field_constraints=fc_list,
            )
            messages.append(spec)
            logger.debug(
                "Loaded message spec: %s (ID=0x%X  required=%s  SA=%s  "
                "cycle_ms=%s  constraints=%d)",
                msg_name, frame_id, required, expected_sa,
                cycle_ms, len(fc_list),
            )

        return messages
